
import time, datetime
from attr import attr, attrib
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select

class Causa:
    def __init__(self, fila, rit, tribunal, acceso, rut, clave, fur):
        self.fila = fila
        self.rit = rit
        self.tribunal = tribunal
        self.acceso = acceso
        self.rut = rut
        self.clave = clave
        self.fur = fur

def login(usuario, pwd):
    driver.get("https://oficinajudicialvirtual.pjud.cl/home/index.php")
    
    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CLASS_NAME, 'dropbtn')))
    driver.find_element(By.CLASS_NAME, "dropbtn").click()
    driver.find_element(By.XPATH, "//div[@id='myDropdown']/a").click()
    driver.find_element(By.ID, "uname").send_keys(usuario)
    driver.find_element(By.ID, "pword").send_keys(pwd)
    driver.find_element(By.ID, "login-submit").click()
    time.sleep(4)

    # Verifica si la clave fue incorrecta
    try:
        driver.find_element(By.CLASS_NAME, 'gob-response-error')
    except:
        pass
    else:
        print('Error')
        return(False)

    # Cierra el modal de bienvenida si aparece
    try:
        driver.find_element(By.ID, 'btnEntendidoBienvenida').click()
        time.sleep(1)
    except:
        pass

    driver.find_element(By.CLASS_NAME, "fa-university").click()
    time.sleep(2)

    # Cierra el modal informativo si aparece
    try:
        driver.find_element(By.ID, 'btnEntendidoMisCausas').click()
        time.sleep(1)
    except:
        pass

    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, 'familiaTab')))
    driver.find_element(By.ID, 'familiaTab').click()
    time.sleep(1)

    driver.find_elements(By.CLASS_NAME, "switch")[6].click()
    time.sleep(3)

    # Selecciona todos los tipos de causas
    estadoCausas = Select(driver.find_element(By.ID, 'estadoCausaMisCauFam'))
    for i in range(12):
        estadoCausas.select_by_index(str(i))
    
    return(True)

def logout():
    driver.find_element(By.XPATH, "//a[@onclick='salir();']").click()
    time.sleep(2)

def nom_tribunal(t):
    res = t
    if (t[0] == '1'):
        res = '1º JUZGADO DE FAMILIA SANTIAGO'
    if (t[0] == '2'):
        res = '2º JUZGADO DE FAMILIA SANTIAGO'
    if (t[0] == '3'):
        res = '3º JUZGADO DE FAMILIA SANTIAGO'
    if (t[0] == '4'):
        res = '4º JUZGADO DE FAMILIA SANTIAGO'
    if (t[0] == 'C'):
        res = 'CENTRO DE MEDIDAS CAUTELARES DE SANTIAGO'
    return(res)

def busca_causa(causa):
    partes = causa.rit.split('-')
    # Ingresa la causa a buscar
    Select(driver.find_element(By.ID, 'tipoMisCauFam')).select_by_value(partes[0])
    driver.find_element(By.ID, 'rolMisCauFam').clear()
    driver.find_element(By.ID, 'rolMisCauFam').send_keys(partes[1])
    driver.find_element(By.ID, 'anhoMisCauFam').clear()
    driver.find_element(By.ID, 'anhoMisCauFam').send_keys(partes[2])
    driver.find_element(By.ID, 'btnConsultaMisCauFam').click()

    # Espera el resultado
    time.sleep(2)

    furPJUD = ''
    revisar = ''
    tramite = ''
    desc = ''

    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, 'dtaTableDetalleMisCauFam')))
    rows = driver.find_elements(By.XPATH, "//table[@id='dtaTableDetalleMisCauFam']/tbody/tr")
    
    for row in rows:
        tds = row.find_elements(By.TAG_NAME, 'td')
        if len(tds) == 7:
            tribunal_en_tabla = nom_tribunal(tds[2].text.strip())
            if tribunal_en_tabla == causa.tribunal:
                tds[0].find_element(By.TAG_NAME, 'a').click()
                WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, 'movimientoFam')))
                cols = driver.find_elements(By.XPATH, "//div[@id='movimientoFam']/div/div/table/tbody/tr/td")
                if len(cols) >= 7:
                    tramite = cols[4].text.strip()
                    desc = cols[5].text.strip()                    
                    fur = cols[6].text.strip()
                    furPJUD = datetime.datetime.strptime(fur, '%d/%m/%Y')
                driver.find_element(By.CLASS_NAME, 'close').click()  
                time.sleep(1)                
    
    if causa.fur != furPJUD:
        revisar = tramite + ', ' + desc
    hoja.cell(causa.fila, 15, furPJUD)
    hoja.cell(causa.fila, 16, revisar)
    guarda_excel()

def revisa_abogado(usuario, pwd, lista):
    if len(lista) == 0:
        return
    login(usuario, pwd)
    for causa in lista:
        busca_causa(causa)
    logout()

def revisa_usuarios(lista):
    if len(lista) == 0:
        return
    for causa in lista:
        res = login(causa.rut, causa.clave)
        if (res):
            busca_causa(causa)
            logout()
        else:
            hoja.cell(causa.fila, 16, 'Clave incorrecta')
            guarda_excel()

def guarda_excel():
    wb.save('Carolina causas.xlsx')

df = pd.read_excel('Carolina causas.xlsx','Asignadas')
por_revisar = df[df['FUR PJUD'].isnull() & df['REVISAR'].isnull()]
activas = por_revisar[por_revisar['T/V'].isin(['D', 'T/R', 'V'])]

print(activas)

caro = []
franco = []
usuarios = []

for i in activas.index:
    causa = Causa(i+2, activas['ROL/RIT'][i], activas['TRIBUNAL'][i], activas['ACCESO'][i], activas['RUT_PATROCINADO'][i], activas['CLAVE'][i], activas['FUR'][i])
    if causa.acceso == 'CC':
        caro.append(causa)
    if causa.acceso == 'FG':
        franco.append(causa)       
    if causa.acceso == 'U':
        usuarios.append(causa)

wb = load_workbook(filename = 'Carolina causas.xlsx')
hoja = wb['Asignadas']

driver = webdriver.Chrome()
driver.maximize_window()

revisa_abogado('138245721', 'POllito-1010', caro)
revisa_abogado('128622497', 'Catalinabarra2614.', franco)
revisa_usuarios(usuarios)

driver.close()


